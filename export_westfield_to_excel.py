import glob
import json
import os
from openpyxl import Workbook


def safe_sheet_name(name: str) -> str:
    # Excel sheet name max length is 31 and cannot contain some chars
    s = name[:31]
    for ch in ('\\', '/', '?', '*', '[', ']', ':'):
        s = s.replace(ch, '_')
    return s


def main():
    cwd = os.getcwd()
    json_files = sorted(glob.glob(os.path.join(cwd, "westfield_stores_*.json")))
    merged_file = os.path.join(cwd, "westfield_stores.json")

    if not json_files and not os.path.exists(merged_file):
        print("No westfield JSON files found (checked for westfield_stores_*.json and westfield_stores.json). Exiting.")
        return

    wb = Workbook()
    # remove default created sheet
    default = wb.active
    wb.remove(default)

    merged_rows = []  # tuples (source, store)

    for jf in json_files:
        try:
            with open(jf, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except Exception as e:
            print(f"Failed to read {jf}: {e}")
            continue
        # Derive a friendly sheet name from filename
        fname = os.path.basename(jf)
        # westfield_stores_<slug>.json -> <slug>
        slug = fname.replace('westfield_stores_', '').replace('.json', '')
        sheet_name = safe_sheet_name(slug)
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["#", "Store Name"])  # header
        for i, store in enumerate(data, start=1):
            ws.append([i, store])
            merged_rows.append((slug, store))
        print(f"Wrote sheet '{sheet_name}' with {len(data)} stores.")

    # If merged grouped display file exists, also add a sheet with that view
    if os.path.exists(merged_file):
        try:
            with open(merged_file, 'r', encoding='utf-8') as f:
                grouped = json.load(f)
        except Exception as e:
            print(f"Failed to read {merged_file}: {e}")
            grouped = None
        if grouped:
            gsheet = wb.create_sheet(title=safe_sheet_name('grouped'))
            gsheet.append(["Line"])  # single-column view
            for i, line in enumerate(grouped, start=1):
                gsheet.append([line])
            print(f"Wrote 'grouped' sheet with {len(grouped)} lines.")

    # Create a merged sheet with source and store, and a deduplicated overview
    ms = wb.create_sheet(title=safe_sheet_name('merged'))
    ms.append(["Source", "Store Name"])  # header
    for src, store in merged_rows:
        ms.append([src, store])

    # Create deduplicated overview sheet
    dedup = []
    seen = set()
    for src, store in merged_rows:
        key = store.strip().lower()
        if key and key not in seen:
            dedup.append((store, src))
            seen.add(key)
    ds = wb.create_sheet(title=safe_sheet_name('unique'))
    ds.append(["Store Name", "First seen (source)"])
    for store, src in dedup:
        ds.append([store, src])
    print(f"Wrote merged sheet with {len(merged_rows)} rows and unique sheet with {len(dedup)} stores.")

    out_path = os.path.join(cwd, 'westfield_stores.xlsx')
    wb.save(out_path)
    print(f"Saved Excel workbook to: {out_path}")


if __name__ == '__main__':
    main()
